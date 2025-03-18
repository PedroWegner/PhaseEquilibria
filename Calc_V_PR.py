from Peng_robinson_module import *
import openpyxl
import os





"""
AQUI É PARA OBTER DE PURO!!!!
wb = openpyxl.load_workbook(
        (os.path.dirname(os.path.abspath(__file__)) + f'\\data\\pentane.xlsx'))

sheet = wb.active
row = 2
for l in filter(None, sheet.iter_rows(min_row=2, values_only=True)):
    T = float(l[0])
    P = float(l[1]) * 1.0e5
    state_liquid.T = T
    state_liquid.P = P
    update_state(state=state_liquid, liq=True)
    V = state_liquid.Z * R * state_liquid.T * 1000 / state_liquid.P
    rho = 1 / V
    erro = abs(float(l[2]) - rho) / float(l[2])

    if erro > 0.35:
        update_state(state=state_liquid, liq=False)
        V = state_liquid.Z * R * state_liquid.T * 1000 / state_liquid.P
        rho = 1 / V
        erro = abs(float(l[2]) - rho) / float(l[2])

    print(f"Z: {state_liquid.Z} - rho: {rho}")
    sheet[f'D{row}'] = V
    sheet[f'E{row}'] = rho
    sheet[f'F{row}'] = state_liquid.Z
    sheet[f'G{row}'] = erro
    row +=1

wb.save((os.path.dirname(os.path.abspath(__file__)) + f'\\data\\teste.xlsx'))
"""

Tc = [190.6, 647.1]
Pc = [45.99e5, 220.55e5]
omega= [0.012, 0.345]
x = [0.9790, 1-0.9790]
T_o = 128.18
P_o = 3.243e5
state_liquid = Peng_robinson_state(
    T=None,
    P=None,
    ncomp=2,
    x=None,
    Tc=Tc,
    Pc=Pc,
    omega=omega
)


# update_state(state=state_liquid, liq=True)
# print(state_liquid.Z)
# V = state_liquid.Z * R * state_liquid.T * 1000 / state_liquid.P
# print(V)
# rho = 1 / V
# print(rho)

wb = openpyxl.load_workbook(
        (os.path.dirname(os.path.abspath(__file__)) + f'\\data\\methane_water.xlsx'))

sheet = wb.active
row = 2
rows = sum(1 for row in sheet.iter_rows(values_only=True) if any(row)) - 1

for l in filter(None, sheet.iter_rows(min_row=2, values_only=True)):
    x_l = l[0].split(";")
    x = []
    x.append(float(x_l[0].replace(",",".")))
    x.append(float(x_l[1].replace(",",".")))
    T = float(l[1])
    P = float(l[2]) * 1.0e5
    state_liquid.x = np.array(x)
    state_liquid.T = T
    state_liquid.P = P
    update_state(state=state_liquid, liq=False)
    V = state_liquid.Z * R * state_liquid.T * 1000 / state_liquid.P
    rho = 1 / V
    erro = abs(float(l[3]) - rho) / float(l[3])
    erro_relativo = (float(l[3]) - rho)**2 / rows
    # if erro > 0.3:
    #     update_state(state=state_liquid, liq=False)
    #     V = state_liquid.Z * R * state_liquid.T * 1000 / state_liquid.P
    #     rho = 1 / V
    #     erro = abs(float(l[3]) - rho) / float(l[3])
    #     erro_relativo = (float(l[3]) - rho) ** 2 / rows

    print(f"Z: {state_liquid.Z} - rho: {rho}")
    sheet[f'E{row}'] = V
    sheet[f'F{row}'] = rho
    sheet[f'G{row}'] = state_liquid.Z
    sheet[f'H{row}'] = erro
    sheet[f'M{row}'] = erro_relativo
    row += 1
wb.save((os.path.dirname(os.path.abspath(__file__)) + f'\\data\\teste.xlsx'))
