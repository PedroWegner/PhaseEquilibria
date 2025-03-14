import numpy as np

from PC_saft_module import *
import openpyxl
import os

# calc_eta(state=state_liquid, P_req=P_o, eta_init=0.3)
# print(state_liquid.eta)
# print(state_liquid.P)
#
# rho_ = (state_liquid.rho / N_Avogrado) * (10**10)**3 * (10**-3)
# print('---')
# print(state_liquid.Z)
# print(1 / rho_)
# print(rho_)
"""
PARA OBTER DOS PUROS
wb = openpyxl.load_workbook(
        (os.path.dirname(os.path.abspath(__file__)) + f'\\data\\teste.xlsx'))

sheet = wb.active
row = 2
for l in filter(None, sheet.iter_rows(min_row=2, values_only=True)):
    T = float(l[0])
    P = float(l[1]) * 1.0e5
    state_liquid.T = T
    state_liquid.P = P
    calc_eta(state=state_liquid, P_req=P, eta_init=0.5)
    rho = (state_liquid.rho / N_Avogrado) * (10**10)**3 * (10**-3)
    V = 1 / rho
    erro = abs(float(l[2]) - rho) / float(l[2])
    if erro > 0.25:
        calc_eta(state=state_liquid, P_req=P, eta_init=0.3)
        rho = (state_liquid.rho / N_Avogrado) * (10 ** 10) ** 3 * (10 ** -3)
        V = 1 / rho
        erro = abs(float(l[2]) - rho) / float(l[2])

    print(f"Z: {state_liquid.Z} - rho: {rho}, V: {V}")
    sheet[f'H{row}'] = V[0]
    sheet[f'I{row}'] = rho[0]
    sheet[f'J{row}'] = state_liquid.Z[0]
    sheet[f'K{row}'] = erro[0]
    row +=1

wb.save((os.path.dirname(os.path.abspath(__file__)) + f'\\data\\teste_2.xlsx'))"""

m = [1.00, 2.0729]
sigma = [3.7039, 2.7852]
epsilon_ = [150.03, 169.21]
T_o = 390.256
P_o = 40.858e5
state_liquid = PC_saft_state(
    T=None,
    P=None,
    ncomp=2,
    x=None,
    m=m,
    sigma=sigma,
    epsilon=epsilon_
)

wb = openpyxl.load_workbook(
        (os.path.dirname(os.path.abspath(__file__)) + f'\\data\\teste.xlsx'))

sheet = wb.active
row = 2
k_ij = np.array([[0.0, 0.065], [0.065, 0.0]])
print(k_ij)
soma = 0.0
rows = sum(1 for row in sheet.iter_rows(values_only=True) if any(row)) - 1
for l in filter(None, sheet.iter_rows(min_row=2, values_only=True)):
    x_l = l[0].split(";")
    x = []
    x.append(float(x_l[0].replace(",",".")))
    x.append(float(x_l[1].replace(",",".")))
    state_liquid.k_ij = k_ij
    T = float(l[1])
    P = float(l[2]) * 1.0e5
    state_liquid.x = np.array(x)
    state_liquid.T = T
    state_liquid.P = P
    calc_eta(state=state_liquid, P_req=P, eta_init=0.3)
    rho = (state_liquid.rho / N_Avogrado) * (10 ** 10) ** 3 * (10 ** -3)
    V = 1 / rho
    erro = abs(float(l[3]) - rho) / float(l[3])
    erro_relativo = (float(l[3]) - rho) ** 2 / rows

    if erro > 0.3:
        calc_eta(state=state_liquid, P_req=P, eta_init=10**-10)
        rho = (state_liquid.rho / N_Avogrado) * (10 ** 10) ** 3 * (10 ** -3)
        V = 1 / rho
        erro = abs(float(l[3]) - rho) / float(l[3])
        erro_relativo = (float(l[3]) - rho) ** 2 / rows

    print(f"Z: {state_liquid.Z} - rho: {rho}")
    sheet[f'I{row}'] = V[0]
    sheet[f'J{row}'] = rho[0]
    sheet[f'K{row}'] = state_liquid.Z[0]
    sheet[f'L{row}'] = erro[0]
    sheet[f'N{row}'] = erro_relativo[0]
    row += 1
    soma += erro

wb.save((os.path.dirname(os.path.abspath(__file__)) + f'\\data\\teste_2.xlsx'))
print(soma)